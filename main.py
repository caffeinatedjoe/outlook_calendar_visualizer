import requests
from ics import Calendar
import json
from datetime import date, timedelta
import google.generativeai as genai
import os
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import config

# --- Configuration ---
EMPLOYEES_FILE = "employees.json"
ZSCALER_CERT_PATH = "zscaler_root.crt" # Path to your Zscaler root certificate

genai.configure(api_key=config.API_KEY)

def load_employees(file_path):
    """
    Loads the hierarchical list of employees from the JSON file.
    Returns the raw hierarchical data.
    """
    try:
        with open(file_path, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Error: Employee file not found at {file_path}")
        return []
    except json.JSONDecodeError:
        print(f"Error: Could not decode JSON from {file_path}")
        return []

def flatten_employees(employees):
    """
    Flattens the hierarchical employee list for LLM processing.
    """
    flat_list = []
    def parse_reports(report_list):
        for person in report_list:
            flat_list.append({
                "name": person["name"],
                "location": person["location"]
            })
            if person.get("reports"):
                parse_reports(person["reports"])
    parse_reports(employees)
    return flat_list

def load_calendar_events(url, event_type):
    """Fetches and parses calendar events from an ICS URL, tagging them with a type."""
    try:
        if not os.path.exists(ZSCALER_CERT_PATH):
            print(f"Error: Zscaler certificate not found at '{ZSCALER_CERT_PATH}'.")
            return []
        response = requests.get(url, verify=ZSCALER_CERT_PATH)
        response.raise_for_status()
    except requests.exceptions.SSLError as e:
        print(f"SSL Error fetching calendar from {url}: {e}")
        return []
    except requests.exceptions.RequestException as e:
        print(f"Error fetching calendar from {url}: {e}")
        return []

    calendar = Calendar(response.text)
    events = []
    for event in calendar.events:
        if event.name:
            events.append({
                "name": event.name,
                "start_date": event.begin.date(),
                "end_date": event.end.date(),
                "type": event_type
            })
    return events


def get_employee_event_mappings_from_llm(unique_event_names, employees):
    """
    Uses the Google Generative AI model to map unique event names to employees.
    """
    model = genai.GenerativeModel('gemini-2.5-flash')
    employee_names = [e['name'] for e in employees]
    
    prompt = f"""
    Analyze the following list of calendar event titles and determine which employee(s) from the provided employee list each event belongs to.
    **Employee List:** {employee_names}
    **Event Titles:** {unique_event_names}
    **Instructions:**
    - Your response MUST be a valid JSON object.
    - The keys of the JSON object should be the event titles.
    - The values should be a list of corresponding employee names.
    - **Holiday Handling:**
        - If a title indicates a US holiday, map it to `["_HOLIDAY_US"]`.
        - If a title indicates a France holiday, map it to `["_HOLIDAY_FRANCE"]`.
        - If a title indicates a company-wide holiday, map it to `["_HOLIDAY_COMPANY"]`.
        - If both countries recognize the holiday, map it to `["_HOLIDAY_COMPANY"]`.
        - Use known national holiday calendars to determine whether the holiday is celebrated in the US, France, or both.
        - If the title suggests a holiday but it is unclear which country(s) it belongs to, make a best-effort guess based on the name.
    - If a title doesn't correspond to any employee or holiday, use an empty list `[]`.
    - Do not include any text or markdown formatting outside the JSON object.
    **Example Response Format:**
    ```json
    {{
      "PTO: Alice": ["Alice"],
      "US Holiday (Thanksgiving) - US Office Closure": ["_HOLIDAY_US"]
    }}
    ```
    """

    try:
        print("Sending prompt to the LLM...")
        response = model.generate_content(prompt)
        cleaned_json = response.text.strip().replace("```json", "").replace("```", "").strip()
        event_to_employee_map = json.loads(cleaned_json)
        print("Successfully received and parsed mappings from LLM.")
        return event_to_employee_map
    except Exception as e:
        print(f"An error occurred during the LLM interaction: {e}")
        return {}


def generate_excel(processed_data, employees, start_date, end_date, file_path):
    """Generates a formatted Excel file from the processed event data."""
    print(f"\nGenerating Excel file at {file_path}...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Calendar View"

    # --- Styles ---
    pto_fill = PatternFill("solid", fgColor="F28C28") # Orange
    travel_fill = PatternFill("solid", fgColor="0070C0") # Blue
    holiday_fill = PatternFill("solid", fgColor="C0C0C0") # Grey
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center")

    # --- Header Creation ---
    all_dates = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]
    weekdays = [d for d in all_dates if d.weekday() < 5]

    # Group dates by month and week
    dates_by_month_week = defaultdict(lambda: defaultdict(list))
    for d in weekdays:
        dates_by_month_week[d.strftime("%B %Y")][d.strftime("W%U")].append(d)

    # Row 1: Month, Row 2: Week, Row 3: Day
    ws.cell(row=3, column=1, value="Employee")
    ws.cell(row=3, column=1).border = border
    current_col = 2

    for month, weeks in dates_by_month_week.items():
        month_start_col = current_col
        for week, days in weeks.items():
            week_start_col = current_col
            for day in days:
                ws.cell(row=3, column=current_col, value=day.strftime("%a %d"))
                ws.cell(row=3, column=current_col).alignment = center_align
                ws.cell(row=3, column=current_col).border = border
                current_col += 1
            
            if len(days) > 0:
                ws.merge_cells(start_row=2, start_column=week_start_col, end_row=2, end_column=current_col - 1)
                ws.cell(row=2, column=week_start_col, value=week)
                ws.cell(row=2, column=week_start_col).alignment = center_align
                ws.cell(row=2, column=week_start_col).border = border

        if (current_col - 1) >= month_start_col:
            ws.merge_cells(start_row=1, start_column=month_start_col, end_row=1, end_column=current_col - 1)
            ws.cell(row=1, column=month_start_col, value=month)
            ws.cell(row=1, column=month_start_col).alignment = center_align
            ws.cell(row=1, column=month_start_col).border = border

    # --- Employee Data ---
    current_row = 4
    def write_employee_row(employee, indent_level):
        nonlocal current_row
        cell = ws.cell(row=current_row, column=1, value=employee['name'])
        cell.border = border
        cell.alignment = Alignment(indent=indent_level)
        
        employee_events = processed_data.get(employee['name'], {})
        
        col_idx = 2
        for report_date in weekdays:
            event_type = employee_events.get(report_date, "")
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = border
            if event_type == "P":
                cell.fill = pto_fill
            elif event_type == "T":
                cell.fill = travel_fill
            elif event_type == "H":
                cell.fill = holiday_fill
            col_idx += 1
        
        current_row += 1
        
        if "reports" in employee:
            for report in employee["reports"]:
                write_employee_row(report, indent_level + 1)

    for employee in employees:
        write_employee_row(employee, 0)

    # --- Formatting ---
    ws.column_dimensions['A'].width = 20
    for i in range(2, current_col):
        ws.column_dimensions[get_column_letter(i)].width = 7

    wb.save(file_path)
    print(f"Successfully generated {file_path}")


def main():
    parser = argparse.ArgumentParser(description="Generate a calendar visualization.")
    parser.add_argument("months", type=int, nargs='?', default=6,
                        help="The number of months to include in the report (default: 6).")
    args = parser.parse_args()

    print("Calendar Visualizer Script")
    
    # 1. Load Employees
    hierarchical_employees = load_employees(EMPLOYEES_FILE)
    if not hierarchical_employees:
        print("No employees loaded, exiting.")
        exit()
    employees = flatten_employees(hierarchical_employees)
    print(f"Loaded {len(employees)} employees.")

    # 2. Load Calendar Events
    print("Loading calendar events...")
    pto_events = load_calendar_events(config.PTO_CALENDAR_URL, "P")
    travel_events = load_calendar_events(config.TRAVEL_CALENDAR_URL, "T")
    all_events = pto_events + travel_events
    print(f"Loaded {len(all_events)} total event entries.")

    # 3. Define the date range for the report
    today = date.today()
    start_date = today.replace(day=1)
    end_date = start_date + timedelta(days=args.months * 30) # Approximation

    # 4. Filter events to the specified date range
    events_in_range = [
        event for event in all_events
        if event['start_date'] < end_date and event['end_date'] > start_date
    ]
    print(f"\nFiltered down to {len(events_in_range)} events within the next {args.months} months.")

    # 5. Get unique event names and map them to employees
    if not events_in_range:
        print("No events to process in the specified date range.")
    else:
        unique_event_names = list(set(event['name'] for event in events_in_range))
        event_to_employee_map = get_employee_event_mappings_from_llm(unique_event_names, employees)

        # 6. Associate events with employees using the map
        print("\nAssociating events with employees...")
        unmatched_event_names = set()
        for event in events_in_range:
            mapped_employees = event_to_employee_map.get(event['name'])
            if mapped_employees is not None:
                event['employee'] = mapped_employees
            else:
                event['employee'] = []
                unmatched_event_names.add(event['name'])
        
        if unmatched_event_names:
            print(f"WARNING: LLM did not return a mapping for: {list(unmatched_event_names)}")

        # 7. Populate new data structure
        for emp in employees:
            emp['PTO'] = []
            emp['Travel'] = []
            emp['Holiday'] = []
        
        employee_map = {emp['name']: emp for emp in employees}
        
        for event in events_in_range:
            if not event.get('employee'):
                continue
            
            event_dates = [event['start_date'] + timedelta(days=i) for i in range((event['end_date'] - event['start_date']).days)]

            for name_or_tag in event['employee']:
                if name_or_tag == "_HOLIDAY_US":
                    for emp in employees:
                        if emp['location'] == 'US': emp['Holiday'].extend(event_dates)
                elif name_or_tag == "_HOLIDAY_FRANCE":
                    for emp in employees:
                        if emp['location'] == 'France': emp['Holiday'].extend(event_dates)
                elif name_or_tag == "_HOLIDAY_COMPANY":
                    for emp in employees: emp['Holiday'].extend(event_dates)
                elif name_or_tag in employee_map:
                    if event['type'] == 'P': employee_map[name_or_tag]['PTO'].extend(event_dates)
                    elif event['type'] == 'T': employee_map[name_or_tag]['Travel'].extend(event_dates)

        # 8. Process the new structure for Excel generation
        print("\nProcessing data for Excel generation...")
        processed_data = {emp['name']: {} for emp in employees}
        for emp in employees:
            for pto_date in emp['PTO']: processed_data[emp['name']][pto_date] = 'P'
            for travel_date in emp['Travel']: processed_data[emp['name']][travel_date] = 'T'
            for holiday_date in emp['Holiday']: processed_data[emp['name']][holiday_date] = 'H'
        
        print("Processing complete.")

        # 9. Generate the Excel output
        output_filename = f"calendar_view_{today.strftime('%m%d%y')}.xlsx"
        generate_excel(processed_data, hierarchical_employees, start_date, end_date, output_filename)

    print("\nScript finished.")

if __name__ == "__main__":
    main()
